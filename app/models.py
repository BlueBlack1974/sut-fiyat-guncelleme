import pandas as pd
import json
import os
import re
from datetime import datetime
from io import BytesIO
from openpyxl import load_workbook

def read_and_process_sut_files(file_stream):
    try:
        df_header = pd.read_excel(file_stream, header=None, nrows=5)
        header_row = None
        
        # EK-2B/2C için başlık kontrolü (debug çıktınıza göre)
        for idx, row in df_header.iterrows():
            row_str = '|'.join([str(cell).strip().upper() for cell in row if pd.notna(cell)])
            if 'İŞLEM KODU' in row_str and 'İŞLEM PUANI' in row_str:
                header_row = idx
                break  # Başlık bulunduğunda döngüyü durdur
        
        # EK-2A kontrolü (gerekirse)
        if header_row is None:
            for idx, row in df_header.iterrows():
                row_str = '|'.join([str(cell).strip().upper() for cell in row if pd.notna(cell)])
                if 'UZMANLIK DALLARI' in row_str and 'ÖH' in row_str and 'ÖTM' in row_str:
                    header_row = idx
                    break  # Başlık bulunduğunda döngüyü durdur
        
        if header_row is None:
            raise ValueError("Excel dosyasında başlık satırı bulunamadı!")

        # DataFrame'i doğru başlıkla oku
        file_stream.seek(0)
        df = pd.read_excel(file_stream, header=header_row)
        df.columns = [str(col).strip().upper().replace(' ', '') for col in df.columns]  # Normalize column names
        
        # Dosya türünü belirle
        if 'İŞLEMKODU' in df.columns and 'İŞLEMPUANI' in df.columns:
            return process_ek2b_or_ek2c(df, 'ek2b' if 'EK-2B' in df.columns else 'ek2c')
        elif 'UZMANLIKDALLARI' in df.columns and 'ÖH' in df.columns and 'ÖTM' in df.columns:
            return process_ek2a(df)
        else:
            raise ValueError("Geçersiz dosya türü. EK-2A, EK-2B veya EK-2C dosyası bekleniyor.")

    except Exception as e:
        print(f"[HATA] Dosya okuma hatası: {str(e)}")
        raise

def process_ek2b_or_ek2c(df, liste_turu):
    print("[DEBUG] DataFrame sütunları:", df.columns.tolist())  # Sütunları göster
    try:
        """EK-2B ve EK-2C için ortak işleme"""
        required_columns = ['İŞLEMKODU', 'İŞLEMPUANI']  # Normalize column names
        for col in required_columns:
            if col not in df.columns:
                raise ValueError(f"{col} kolonu eksik!")

        data = []
        for _, row in df.iterrows():
            islem_kodu = str(row['İŞLEMKODU']).strip()
            if not islem_kodu or islem_kodu.lower() in ['nan', 'none', '']:
                continue  # islem_kodu boş veya geçersizse atla
            
            # Ortadan çizgili olanları atla
            if '̶' in islem_kodu:
                continue
            
            # İşlem kodunda uzun yazı bulunanları atla
            if len(islem_kodu) > 10:
                continue
            
            puan_str = str(row['İŞLEMPUANI']).replace(',', '.').replace(' ', '')
            print(f"[DEBUG] İşlem Kodu: {islem_kodu}, Puan: {puan_str}")  # Debug
        
            try:
                islem_puani = float(puan_str)
            except ValueError:
                print(f"Geçersiz puan: {islem_kodu}")
                continue

            # Hesaplamalar
            kdv_haric = round(islem_puani * 0.593, 2)
            kdv_dahil = round(kdv_haric * 1.1, 4)  # 4 haneye yuvarla

            data.append({
                'islem_kodu': islem_kodu,
                'kdv_haric_fiyat': kdv_haric,
                'kdv_dahil_fiyat': kdv_dahil,
                'liste_turu': liste_turu
            })
        print(f"[DEBUG] İşlenen veri sayısı: {len(data)}")  # Veri sayısı
        return data
    except Exception as e:
        print("[HATA] JSON kayıt hatası:", str(e))
        raise

def process_ek2a(df):
    """EK-2A işleme"""
    required_columns = ['UZMANLIKDALLARI', 'ÖH', 'ÖTM']  # Normalize column names
    for col in required_columns:
        if col not in df.columns:
            raise ValueError(f"{col} kolonu eksik!")

    data = []
    for _, row in df.iterrows():
        uzmanlik = str(row['UZMANLIKDALLARI']).strip()
        if not uzmanlik or uzmanlik.lower() in ['nan', 'none', '']:
            continue

        # ÖH ve ÖTM değerlerini temizle
        oh = str(row['ÖH']).replace(',', '.').replace(' ', '')
        otm = str(row['ÖTM']).replace(',', '.').replace(' ', '')
        
        try:
            oh_kdv_haric = float(oh)
            otm_kdv_haric = float(otm)
        except ValueError:
            print(f"Geçersiz fiyat: {uzmanlik}")
            continue

        # KDV dahil hesapla
        oh_kdv_dahil = round(oh_kdv_haric * 1.1, 2)
        otm_kdv_dahil = round(otm_kdv_haric * 1.1, 2)

        data.append({
            'uzmanlik_dali': uzmanlik,
            'oh_kdv_haric': oh_kdv_haric,
            'oh_kdv_dahil': oh_kdv_dahil,
            'otm_kdv_haric': otm_kdv_haric,
            'otm_kdv_dahil': otm_kdv_dahil,
            'liste_turu': 'ek2a'
        })
    return data

def save_to_json(data, liste_turu):
    try:
        print(f"[DEBUG] Kaydedilecek veri ({liste_turu}):", data)  # Veriyi göster
        """JSON'a kaydet (Tüm verileri tarih bazlı sakla)"""
        json_path = os.path.join('data', 'sut_fiyatlari.json')
        os.makedirs(os.path.dirname(json_path), exist_ok=True)
        
        # Mevcut verileri oku veya yeni oluştur
        if os.path.exists(json_path):
            with open(json_path, 'r', encoding='utf-8') as f:
                try:
                    existing_data = json.load(f)
                except json.JSONDecodeError:
                    existing_data = {}
        else:
            existing_data = {}
        
        # Yeni veriyi ekle
        current_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        if liste_turu not in existing_data:
            existing_data[liste_turu] = {}
        existing_data[liste_turu][current_time] = data
        
        # Dosyaya yaz
        with open(json_path, 'w', encoding='utf-8') as f:
            json.dump(existing_data, f, ensure_ascii=False, indent=4)
    except Exception as e:
        print("[HATA] JSON kayıt hatası:", str(e))
        raise

def is_matching_description(description, specialist):
    """İki metin arasında eşleşme kontrolü yapar"""
    print(f"\nKarşılaştırma yapılıyor:")
    print(f"Excel'deki metin: '{description}'")
    print(f"Uzmanlık dalı: '{specialist}'")
    
    # Küçük harfe çevir
    description = description.lower()
    specialist = specialist.lower()
    
    print(f"Küçük harfli: '{description}' - '{specialist}'")
    
    # Eş anlamlı terimler
    synonyms = {
        'iç hastalıkları': ['dahiliye'],
        'dahiliye': ['iç hastalıkları'],
        'kulak burun boğaz': ['kbb'],
        'kbb': ['kulak burun boğaz'],
        'fizik tedavi': ['fizik tedavi ve rehabilitasyon', 'ftr'],
        'ftr': ['fizik tedavi', 'fizik tedavi ve rehabilitasyon'],
        'kadın hastalıkları': ['kadın doğum', 'kadın hastalıkları ve doğum', 'jinekoloji'],
        'çocuk hastalıkları': ['pediatri'],
        'pediatri': ['çocuk hastalıkları'],
        'ruh hastalıkları': ['psikiyatri'],
        'psikiyatri': ['ruh hastalıkları']
    }
    
    # Parantezleri ve özel kelimeleri kaldır
    ignore_patterns = [
        r'\[.*?\]',      # Köşeli parantez içi
        r'\(.*?\)',      # Normal parantez içi
        r'\bpaket\b',    # "paket" kelimesi
        r'\bmuayenesi\b' # "muayenesi" kelimesi
    ]
    
    # Karşılaştırma için metni temizle
    temp_description = description
    for pattern in ignore_patterns:
        temp_description = re.sub(pattern, '', temp_description)
    
    # Boşlukları temizle
    temp_description = ' '.join(temp_description.split())
    specialist = ' '.join(specialist.split())
    
    print(f"Temizlenmiş metin: '{temp_description}'")
    print(f"Aranan uzman: '{specialist}'")
    
    # Tam eşleşme kontrolü
    if specialist in temp_description:
        print("Tam eşleşme bulundu!")
        return True, 1.0
    
    # Eş anlamlı kelime kontrolü
    if specialist in synonyms:
        print(f"Eş anlamlılar kontrol ediliyor: {synonyms[specialist]}")
        for synonym in synonyms[specialist]:
            if synonym in temp_description:
                print(f"Eş anlamlı kelime bulundu: {synonym}")
                return True, 1.0
    
    # Kelime bazlı kontrol
    specialist_words = set(specialist.split())
    description_words = set(temp_description.split())
    
    print(f"Uzman kelimeleri: {specialist_words}")
    print(f"Metin kelimeleri: {description_words}")
    
    # Eş anlamlı kelimeleri de ekle
    expanded_description_words = set(description_words)
    for word in description_words:
        if word in synonyms:
            for synonym in synonyms[word]:
                expanded_description_words.update(synonym.split())
    
    print(f"Genişletilmiş kelimeler: {expanded_description_words}")
    
    # Ortak kelime sayısını bul
    common_words = specialist_words.intersection(expanded_description_words)
    
    print(f"Ortak kelimeler: {common_words}")
    
    # Benzerlik oranını hesapla
    if len(specialist_words) > 0:
        match_ratio = len(common_words) / len(specialist_words)
        print(f"Benzerlik oranı: %{match_ratio*100:.1f}")
        # En az %80 benzerlik varsa eşleşme kabul et
        if match_ratio >= 0.8:
            return True, match_ratio
            
    return False, 0.0

def update_excel_from_json(file_stream, code_column_letter, description_column_letter, price_column_letter, price_type='dahil', hospital_type='oh'):
    """Excel dosyasını JSON'daki fiyatlarla günceller"""
    try:
        print("\nExcel dosyası yükleniyor...")
        # Excel dosyasını bellekte aç
        workbook = load_workbook(file_stream)
        sheet = workbook.active
        
        print("\nJSON verisi yükleniyor...")
        # JSON verilerini oku
        json_file = os.path.join('data', 'sut_fiyatlari.json')
        with open(json_file, 'r', encoding='utf-8') as f:
            try:
                sut_data = json.load(f)
            except json.JSONDecodeError:
                raise ValueError("Geçersiz JSON formatı")

        # Kolon indekslerini hesapla
        code_col_idx = column_letter_to_index(code_column_letter)
        desc_col_idx = column_letter_to_index(description_column_letter)
        price_col_idx = column_letter_to_index(price_column_letter)
        
        # Güncelleme işlemi
        updated_rows = 0
        not_found_codes = []
        total_rows = sheet.max_row

        # EK-2B ve EK-2C için en son fiyatları al
        latest_prices = {}
        for liste_turu in ['ek2b', 'ek2c']:
            if liste_turu in sut_data and len(sut_data[liste_turu]) > 0:
                latest_date = max(sut_data[liste_turu].keys())
                for item in sut_data[liste_turu][latest_date]:
                    key = str(item['islem_kodu']).strip()
                    if key:  # Boş işlem kodlarını atla
                        latest_prices[key] = item[f'kdv_{price_type}_fiyat']

        # EK-2A için en son fiyatları al
        specialist_prices = {}
        if 'ek2a' in sut_data and sut_data['ek2a']:
            latest_date = max(sut_data['ek2a'].keys())
            latest_data = sut_data['ek2a'][latest_date]
            for item in latest_data:
                if 'uzmanlik_dali' in item and 'liste_turu' in item and item['liste_turu'] == 'ek2a':
                    # Hastane türü ve fiyat tipine göre fiyatı al
                    if price_type == 'dahil':
                        price = item[f'{hospital_type}_kdv_dahil']
                    else:
                        price = item[f'{hospital_type}_kdv_haric']
                    specialist_prices[item['uzmanlik_dali']] = price

        print("\nFiyatlar güncelleniyor...")
        for row in range(2, total_rows + 1):  # İlk satır başlık
            try:
                # İşlem kodunu al
                code = str(sheet.cell(row=row, column=code_col_idx + 1).value).strip()
                if not code:  # Boş satırları atla
                    continue

                updated = False
                
                # Önce normal kod kontrolü (EK-2B ve EK-2C)
                if code in latest_prices:
                    old_price = sheet.cell(row=row, column=price_col_idx + 1).value
                    new_price = latest_prices[code]
                    sheet.cell(row=row, column=price_col_idx + 1, value=new_price)
                    updated_rows += 1
                    updated = True
                    if row < 5:  # İlk 3 güncellemeyi göster
                        print(f"Güncelleme (EK-2B/C): Kod={code}, Eski={old_price}, Yeni={new_price}")
                
                # P520030 kodu için uzmanlık dalı kontrolü (EK-2A)
                elif code == "P520030" and specialist_prices:
                    description = str(sheet.cell(row=row, column=desc_col_idx + 1).value).strip()
                    print(f"\nP520030 kodu bulundu: {description}")
                    
                    # Uzmanlık dalını bul
                    found_specialist = None
                    max_match_ratio = 0
                    
                    # Her bir uzmanlık dalı için eşleşme kontrolü
                    for specialist in specialist_prices:
                        is_match, match_ratio = is_matching_description(description, specialist)
                        if is_match and match_ratio > max_match_ratio:
                            max_match_ratio = match_ratio
                            found_specialist = specialist
                    
                    # Fiyatı güncelle
                    if found_specialist:
                        old_price = sheet.cell(row=row, column=price_col_idx + 1).value
                        new_price = specialist_prices[found_specialist]
                        sheet.cell(row=row, column=price_col_idx + 1, value=new_price)
                        updated_rows += 1
                        updated = True
                        if row < 5:  # İlk 3 güncellemeyi göster
                            print(f"Güncelleme (EK-2A): Kod={code}, Uzman={found_specialist}, Eski={old_price}, Yeni={new_price}")
                            if max_match_ratio < 1:
                                print(f"  - Benzerlik: %{max_match_ratio*100:.1f}")
                                print(f"  - Excel'deki metin: {description}")
                
                if not updated:
                    not_found_codes.append(code)
                    if len(not_found_codes) < 4:  # İlk 3 bulunamayan kodu göster
                        if code == "P520030":
                            print(f"Bulunamayan uzmanlık dalı: {description}")
                        else:
                            print(f"Bulunamayan kod: {code}")
                        
            except Exception as e:
                print(f"HATA: Satır {row} güncellenirken hata: {str(e)}")
                continue

        # Güncellenmiş dosyayı bellekte tut
        output = BytesIO()
        workbook.save(output)
        output.seek(0)
        
        # Workbook'u kapat
        workbook.close()
        
        # Sonuçları göster
        print("\nGüncelleme Tamamlandı!")
        print(f"Toplam {total_rows} satırdan {updated_rows} satır güncellendi.")
        print(f"Bulunamayan kod sayısı: {len(not_found_codes)}")
        
        return {
            'success': True,
            'message': 'Güncelleme başarılı',
            'total_rows': total_rows,
            'updated_rows': updated_rows,
            'not_found_rows': len(not_found_codes),
            'file_stream': output
        }

    except Exception as e:
        print(f"\nHATA: {str(e)}")
        return {
            'success': False,
            'message': str(e),
            'total_rows': 0,
            'updated_rows': 0,
            'not_found_rows': 0,
            'file_stream': None
        }

def column_letter_to_index(letter):
    """Excel kolon harfini indekse çevirir (A=0, B=1, ...)"""
    letter = letter.upper().strip()
    index = 0
    for char in letter:
        index = index * 26 + (ord(char) - ord('A') + 1)
    return index - 1