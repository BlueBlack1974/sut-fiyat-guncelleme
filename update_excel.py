import os
import json
import openpyxl
from openpyxl.utils import column_index_from_string

def update_excel_from_json(excel_path, code_column, price_column, price_type='dahil'):
    """Excel dosyasını JSON'daki fiyatlarla günceller (Sadece EK-2B ve EK-2C için)"""
    try:
        print(f"\nExcel dosyası yükleniyor: {excel_path}")
        # Excel dosyasını aç
        workbook = openpyxl.load_workbook(excel_path)
        sheet = workbook.active
        
        print("\nJSON verisi yükleniyor...")
        # JSON verilerini oku
        json_file = os.path.join('data', 'sut_fiyatlari.json')
        with open(json_file, 'r', encoding='utf-8') as f:
            try:
                sut_data = json.load(f)
            except json.JSONDecodeError:
                raise ValueError("Geçersiz JSON formatı")

        # EK-2B ve EK-2C için en son fiyatları al
        latest_prices = {}
        for liste_turu in ['ek2b', 'ek2c']:
            if liste_turu in sut_data and len(sut_data[liste_turu]) > 0:
                latest_date = max(sut_data[liste_turu].keys())
                for item in sut_data[liste_turu][latest_date]:
                    key = str(item['islem_kodu']).strip()
                    if key:  # Boş işlem kodlarını atla
                        latest_prices[key] = item[f'kdv_{price_type}_fiyat']

        if not latest_prices:
            raise ValueError("EK-2B veya EK-2C için fiyat verisi bulunamadı!")

        print(f"JSON'dan yüklenen fiyat sayısı: {len(latest_prices)}")
        
        # Kolon indekslerini hesapla
        code_col_idx = column_index_from_string(code_column.upper())
        price_col_idx = column_index_from_string(price_column.upper())

        print(f"İşlem kodu kolonu: {code_column} (index: {code_col_idx})")
        print(f"Fiyat kolonu: {price_column} (index: {price_col_idx})")

        # Güncelleme işlemi
        updated_rows = 0
        not_found_codes = []
        total_rows = sheet.max_row
        
        print("\nFiyatlar güncelleniyor...")
        for row in range(2, total_rows + 1):  # İlk satır başlık olduğu için 2'den başla
            try:
                # İşlem kodunu al ve temizle
                current_key = str(sheet.cell(row=row, column=code_col_idx).value).strip()
                
                if not current_key:  # Boş satırları atla
                    continue
                    
                # JSON'da fiyat varsa güncelle
                if current_key in latest_prices:
                    old_price = sheet.cell(row=row, column=price_col_idx).value
                    new_price = latest_prices[current_key]
                    sheet.cell(row=row, column=price_col_idx, value=new_price)
                    updated_rows += 1
                    if row < 5:  # İlk 3 güncellemeyi göster
                        print(f"Güncelleme: Kod={current_key}, Eski={old_price}, Yeni={new_price}")
                else:
                    not_found_codes.append(current_key)
                    if len(not_found_codes) < 4:  # İlk 3 bulunamayan kodu göster
                        print(f"Bulunamayan kod: {current_key}")
            except Exception as e:
                print(f"HATA: Satır {row} güncellenirken hata: {str(e)}")
                continue

        # Değişiklikleri kaydet
        print("\nDeğişiklikler kaydediliyor...")
        workbook.save(excel_path)
        
        # Sonuçları göster
        print("\nGüncelleme Tamamlandı!")
        print(f"Toplam {total_rows} satırdan {updated_rows} satır güncellendi.")
        print(f"Bulunamayan kod sayısı: {len(not_found_codes)}")
        
        return True

    except Exception as e:
        print(f"\nHATA: {str(e)}")
        return False

def main():
    print("Excel Fiyat Güncelleme Aracı")
    print("-" * 30)
    
    # Kullanıcıdan bilgileri al
    excel_path = input("\nExcel dosyasının tam yolunu girin: ").strip('"')  # Tırnak işaretlerini kaldır
    code_column = input("İşlem kodu kolonunu girin (örn: A, B, C): ").strip()
    price_column = input("Fiyat kolonunu girin (örn: A, B, C): ").strip()
    price_type = input("Fiyat türünü girin (dahil/haric) [varsayılan: dahil]: ").strip() or 'dahil'
    
    # Güncelleme işlemini başlat
    success = update_excel_from_json(excel_path, code_column, price_column, price_type)
    
    if success:
        print("\nİşlem başarıyla tamamlandı!")
    else:
        print("\nİşlem sırasında bir hata oluştu!")
    
    input("\nKapatmak için Enter'a basın...")

if __name__ == "__main__":
    main()
