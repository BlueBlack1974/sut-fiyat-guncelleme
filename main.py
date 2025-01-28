from flask import Flask, render_template, redirect, url_for, flash, jsonify, send_file, after_this_request, request
from flask_wtf import FlaskForm
from wtforms import FileField, RadioField, SubmitField, StringField
from wtforms.validators import DataRequired
from werkzeug.utils import secure_filename
import os
import json
import shutil
import time
from datetime import datetime, timedelta
from io import BytesIO
import openpyxl
import pandas as pd

# Flask app configuration
app = Flask(__name__,
    template_folder='templates',
    static_folder='static'
)

# Config
app.config['SECRET_KEY'] = 'your-secret-key-here'
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16MB max file size
app.config['UPLOAD_FOLDER'] = os.path.join(os.getcwd(), 'uploads')
app.config['TEMP_FOLDER'] = os.path.join(os.getcwd(), 'temp')

# Form classes
class UpdateForm(FlaskForm):
    excel_file = FileField('Excel Dosyası', validators=[DataRequired()])
    code_column = StringField('SUT Kodu Kolonu', validators=[DataRequired()])
    description_column = StringField('İşlem Açıklaması Kolonu', validators=[DataRequired()])
    price_column = StringField('Fiyat Kolonu', validators=[DataRequired()])
    price_type = RadioField('Fiyat Türü',
                          choices=[('dahil', 'KDV Dahil'),
                                 ('haric', 'KDV Hariç')],
                          default='dahil',
                          validators=[DataRequired()])
    hospital_type = RadioField('Hastane Türü',
                             choices=[('ozel_hastane', 'Özel Hastane'),
                                    ('ozel_tip_merkezi', 'Özel Tıp Merkezi')],
                             default='ozel_hastane',
                             validators=[DataRequired()])
    submit = SubmitField('Güncelle')

class AdminForm(FlaskForm):
    sut_file = FileField('SUT Dosyası')  # Dosya yükleme alanı
    submit = SubmitField('SUT Verilerini Güncelle')
    clear_data = SubmitField('Verileri Temizle')  # Temizleme butonu eklendi

# Uzmanlık dalları ve eş anlamlıları
SPECIALTY_MAPPING = {
    'İç hastalıkları': ['Dahiliye', 'Internal'],
    'Kadın hastalıkları ve doğum': ['Jinekoloji', 'Kadın Doğum', 'Jinekolog'],
    'Kulak Burun Boğaz Hastalıkları': ['KBB', 'Kulak-Burun-Boğaz'],
    'Fizik tedavi ve rehabilitasyon': ['FTR', 'Fizik Tedavi'],
    'Göz hastalıkları': ['Oftalmoloji'],
    'Çocuk sağlığı ve hastalıkları': ['Pediatri'],
    'Ruh Sağlığı ve Hastalıkları': ['Psikiyatri'],
    'Deri ve zührevi hastalıkları': ['Dermatoloji', 'Cildiye'],
    'Nöroloji': ['Sinir Hastalıkları'],
    'Ortopedi ve travmatoloji': ['Ortopedi'],
    'Üroloji': ['Bevliye'],
    'Kardiyoloji': ['Kalp Hastalıkları'],
    'Kalp ve Damar Cerrahisi': ['Kardiyovasküler Cerrahi'],
    'Gastroenteroloji': []  # Eş anlamlısı yok, direkt kendisi kullanılacak
}

# Göz ardı edilecek kelimeler
IGNORE_WORDS = ['muayene', 'muayenesi', 'Muayene', 'Muayenesi', 'MUAYENE', 'MUAYENESI',
                'paket', 'Paket', 'PAKET', '[', ']', '[ paket ]', '[paket]', '**']

def clean_specialty_text(text):
    """Metinden göz ardı edilecek kelimeleri çıkarır ve temizler"""
    print(f"\nTemizleme öncesi metin: '{text}'")
    text = text.lower().strip()
    for word in IGNORE_WORDS:
        text = text.replace(word.lower(), '')
    text = ' '.join(text.split())  # Fazla boşlukları temizle
    print(f"Temizleme sonrası metin: '{text}'")
    return text

def find_specialty_match(text, target_specialty):
    """Verilen metin içinde uzmanlık dalı eşleşmesi arar"""
    if not text or not target_specialty:
        return False
        
    # Metinleri temizle
    cleaned_text = clean_specialty_text(text)
    cleaned_target = clean_specialty_text(target_specialty)
    print(f"\nKarşılaştırma: '{cleaned_text}' ile '{cleaned_target}'")
    
    # Tam eşleşme kontrolü
    if cleaned_text == cleaned_target:
        print("Tam eşleşme bulundu!")
        return True
        
    # Ana uzmanlık dalı kontrolü
    if target_specialty in SPECIALTY_MAPPING:
        # Önce ana uzmanlık dalının kendisiyle kontrol
        if cleaned_target in cleaned_text or cleaned_text in cleaned_target:
            print(f"Ana uzmanlık dalı eşleşmesi bulundu!")
            return True
            
        # Eş anlamlıları kontrol et
        for synonym in SPECIALTY_MAPPING[target_specialty]:
            cleaned_synonym = clean_specialty_text(synonym)
            print(f"Eş anlamlı kontrol: '{cleaned_text}' ile '{cleaned_synonym}'")
            
            if cleaned_text in cleaned_synonym or cleaned_synonym in cleaned_text:
                print(f"Eş anlamlı eşleşme bulundu: {synonym}")
                return True
                
    # Benzerlik oranı kontrolü
    from difflib import SequenceMatcher
    similarity = SequenceMatcher(None, cleaned_text, cleaned_target).ratio()
    print(f"Benzerlik oranı: {similarity}")
    if similarity > 0.8:
        print("Benzerlik eşleşmesi bulundu!")
        return True
    
    return False

def find_best_specialty_match(text, specialty_mapping):
    """Verilen metin için en iyi uzmanlık dalı eşleşmesini bul"""
    if not text:
        return None
        
    text = text.lower().strip()
    best_match = None
    max_similarity = 0
    
    for main_specialty, alternatives in specialty_mapping.items():
        # Ana uzmanlık dalı kontrolü
        if main_specialty.lower() in text:
            return main_specialty
            
        # Alternatif isimler kontrolü
        for alt in alternatives:
            if alt.lower() in text:
                return main_specialty
                
        # Kısmi eşleşme kontrolü
        main_words = set(main_specialty.lower().split())
        text_words = set(text.split())
        common_words = main_words.intersection(text_words)
        
        if common_words:
            similarity = len(common_words) / len(main_words)
            if similarity > max_similarity:
                max_similarity = similarity
                best_match = main_specialty
                
    return best_match if max_similarity > 0.5 else None

def update_excel_from_json(excel_file, hospital_type, code_column, description_column, price_column, price_type):
    """Excel dosyasını JSON verilerine göre güncelle"""
    try:
        # JSON verilerini oku
        data_dir = os.path.join(os.getcwd(), 'data')
        json_path = os.path.join(data_dir, 'sut_fiyatlari.json')
        
        if not os.path.exists(json_path):
            # Boş değerler yerine 0 dön
            return None, 0, 0
            
        with open(json_path, 'r', encoding='utf-8') as f:
            json_data = json.load(f)
            
        # SUT verilerini al
        sut_data = {}
        specialty_prices = {}
        
        # Ek2a listesinden uzmanlık dalı fiyatlarını al
        if 'ek2a' in json_data and json_data['ek2a']:  # json_data['ek2a'] boş değilse
            latest_date = max(json_data['ek2a'].keys()) if json_data['ek2a'] else None
            if latest_date and json_data['ek2a'][latest_date]:
                for item in json_data['ek2a'][latest_date]:
                    if isinstance(item, dict) and 'uzmanlik_dali' in item:
                        specialty = item['uzmanlik_dali']
                        if hospital_type == 'ozel_hastane':
                            price = item['oh_kdv_dahil'] if price_type == 'dahil' else item['oh_kdv_haric']
                        else:  # ozel_tip_merkezi
                            price = item['otm_kdv_dahil'] if price_type == 'dahil' else item['otm_kdv_haric']
                        specialty_prices[specialty] = price
                print(f"\nEk2a'dan {len(specialty_prices)} uzmanlık dalı fiyatı yüklendi")
                print("Yüklenen uzmanlık dalları ve fiyatları:")
                for specialty, price in specialty_prices.items():
                    print(f"{specialty}: {price}")
        
        # Ek2b ve Ek2c listelerinden fiyatları al
        for liste_turu in ['ek2b', 'ek2c']:
            if liste_turu in json_data and json_data[liste_turu]:  # json_data[liste_turu] boş değilse
                latest_date = max(json_data[liste_turu].keys()) if json_data[liste_turu] else None
                if latest_date and json_data[liste_turu][latest_date]:
                    print(f"\n{liste_turu.upper()} Listesi:")
                    items = json_data[liste_turu][latest_date]
                    
                    for item in items:
                        if isinstance(item, dict) and 'islem_kodu' in item:
                            code = str(item['islem_kodu']).strip()
                            if code != 'P520030':  # P520030 dışındaki kodlar için
                                if price_type == 'dahil':
                                    fiyat = item['kdv_dahil_fiyat']
                                else:
                                    fiyat = item['kdv_haric_fiyat']
                                sut_data[code] = fiyat
                                print(f"- {code}: {fiyat} TL")
        
        # Excel'i yükle
        try:
            workbook = openpyxl.load_workbook(excel_file, data_only=True, read_only=False)
        except Exception as e:
            try:
                if excel_file.endswith('.xls'):
                    temp_dir = os.path.join(os.getcwd(), 'temp')
                    os.makedirs(temp_dir, exist_ok=True)
                    xlsx_path = os.path.join(temp_dir, 'temp.xlsx')
                    
                    # .xls dosyasını .xlsx'e çevir
                    df = pd.read_excel(excel_file)
                    df.to_excel(xlsx_path, index=False)
                    
                    # Yeni .xlsx dosyasını yükle
                    workbook = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=False)
                    
                    # Geçici dosyayı sil
                    delete_file_after_delay(xlsx_path)
                else:
                    raise e
            except Exception as e2:
                return None, 0, 0
                
        sheet = workbook.active
        
        # Kolon harflerini al
        code_col = code_column.upper()
        desc_col = description_column.upper()
        price_col = price_column.upper()
        
        # Başlık satırını bul
        header_row = 1  # Varsayılan olarak ilk satır
        updated_rows = 0
        total_rows = 0
        not_found_rows = 0
        
        for row in range(header_row + 1, sheet.max_row + 1):
            total_rows += 1
            
            # SUT kodunu al
            code_cell = f"{code_col}{row}"
            code = str(sheet[code_cell].value).strip() if sheet[code_cell].value else ""
            
            if code == "P520030":
                # İşlem açıklamasını al
                desc_cell = f"{desc_col}{row}"
                description = str(sheet[desc_cell].value).strip() if sheet[desc_cell].value else ""
                print(f"\nP520030 için işlem açıklaması kontrolü:")
                print(f"Satır {row} - İşlem açıklaması: '{description}'")
                
                # En uygun uzmanlık dalı eşleşmesini bul
                best_match = None
                best_similarity = 0
                
                for specialty in specialty_prices.keys():
                    print(f"\nUzmanlık dalı kontrolü: '{specialty}'")
                    if find_specialty_match(description, specialty):
                        # Eşleşme bulundu, fiyatı güncelle
                        new_price = specialty_prices[specialty]
                        old_price = sheet[f"{price_col}{row}"].value
                        sheet[f"{price_col}{row}"].value = new_price
                        updated_rows += 1
                        print(f"Eşleşme bulundu!")
                        print(f"Uzmanlık: {specialty}")
                        print(f"Fiyat güncellendi: {old_price} -> {new_price}")
                        break
                else:
                    print(f"Eşleşme bulunamadı!")
                    not_found_rows += 1
            
            elif code in sut_data:
                new_price = sut_data[code]
                old_price = sheet[f"{price_col}{row}"].value
                sheet[f"{price_col}{row}"].value = new_price
                updated_rows += 1
            else:
                not_found_rows += 1
        
        # Excel verilerini hafızada byte dizisine çevir
        excel_buffer = BytesIO()
        workbook.save(excel_buffer)
        excel_buffer.seek(0)
        
        return excel_buffer, updated_rows, not_found_rows
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        return None, 0, 0

def cleanup_old_files():
    """1 saatten eski geçici dosyaları temizle"""
    temp_dir = os.path.join(os.getcwd(), 'temp')
    if not os.path.exists(temp_dir):
        return
        
    try:
        current_time = datetime.now()
        for filename in os.listdir(temp_dir):
            file_path = os.path.join(temp_dir, filename)
            file_modified = datetime.fromtimestamp(os.path.getmtime(file_path))
            if current_time - file_modified > timedelta(hours=1):
                if os.path.isfile(file_path):
                    os.remove(file_path)
                elif os.path.isdir(file_path):
                    shutil.rmtree(file_path)
    except Exception as e:
        print(f"Cleanup error: {str(e)}")

def delete_file_after_delay(file_path, delay=1):
    """Belirtilen süre sonra dosyayı sil"""
    def delete_file():
        try:
            time.sleep(delay)
            if os.path.exists(file_path):
                os.remove(file_path)
        except Exception as e:
            print(f"Error deleting file: {str(e)}")
            
    import threading
    thread = threading.Thread(target=delete_file)
    thread.start()

def check_file_access(file_stream):
    """Excel dosyasının erişilebilir olup olmadığını kontrol et"""
    try:
        workbook = openpyxl.load_workbook(file_stream, read_only=True)
        workbook.close()
        return True
    except Exception as e:
        return False

def read_and_process_sut_files():
    """SUT dosyalarını okur ve işler"""
    try:
        json_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'data', 'sut_fiyatlari.json')
        
        # Eğer dosya varsa, mevcut verileri oku
        if os.path.exists(json_path):
            with open(json_path, 'r', encoding='utf-8') as f:
                return json.load(f)
        
        # Dosya yoksa yeni yapı oluştur
        return {
            'ek2a': {},
            'ek2b': {},
            'ek2c': {}
        }
    except Exception as e:
        print(f"Hata: {str(e)}")
        return None

def save_to_json(data):
    """Verileri JSON dosyasına kaydeder"""
    try:
        json_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'data', 'sut_fiyatlari.json')
        with open(json_path, 'w', encoding='utf-8') as f:
            json.dump(data, f, ensure_ascii=False, indent=4)
        return json_path
    except Exception as e:
        print(f"JSON kaydetme hatası: {str(e)}")
        return None

def process_ek2a(df):
    """Ek2a Excel dosyasını işler"""
    try:
        # Başlık satırı 3. satır (index=2)
        header_row = 2
        
        # Kolonları belirle
        df.columns = df.iloc[header_row]
        df = df.iloc[header_row + 1:]  # Veriler 4. satırdan başlıyor
        
        # Kolon isimlerini bul
        uzmanlik_kolonu = None
        oh_kolonu = None
        otm_kolonu = None
        
        for col in df.columns:
            if not isinstance(col, str):
                continue
                
            col_clean = str(col).upper().strip()
            if col_clean == "UZMANLIK DALLARI":
                uzmanlik_kolonu = col
            elif col_clean == "ÖH":
                oh_kolonu = col
            elif col_clean == "ÖTM":
                otm_kolonu = col
        
        if not uzmanlik_kolonu or (not oh_kolonu and not otm_kolonu):
            raise ValueError("Gerekli kolonlar bulunamadı (UZMANLIK DALLARI ve ÖH/ÖTM)")
        
        # Veriyi işle
        uzmanlik_dallari = []
        for _, row in df.iterrows():
            try:
                uzmanlik = str(row[uzmanlik_kolonu]).strip() if pd.notna(row[uzmanlik_kolonu]) else None
                oh = str(row[oh_kolonu]).strip() if oh_kolonu and pd.notna(row[oh_kolonu]) else None
                otm = str(row[otm_kolonu]).strip() if otm_kolonu and pd.notna(row[otm_kolonu]) else None
                
                # Boş satırları atla
                if not uzmanlik:
                    continue
                    
                # Sayısal değerleri al
                oh_deger = float(oh) if oh and oh not in ["*", ""] and pd.notna(oh) else 0.0
                otm_deger = float(otm) if otm and otm not in ["*", ""] and pd.notna(otm) else 0.0
                
                # KDV hesapla (%10)
                oh_kdv_dahil = round(oh_deger * 1.1, 1) if oh_deger > 0 else 0.0
                otm_kdv_dahil = round(otm_deger * 1.1, 1) if otm_deger > 0 else 0.0
                
                # En az bir değer varsa ekle
                if oh_deger > 0 or otm_deger > 0:
                    uzmanlik_dallari.append({
                        'uzmanlik_dali': uzmanlik,
                        'oh_kdv_haric': oh_deger,
                        'oh_kdv_dahil': oh_kdv_dahil,
                        'otm_kdv_haric': otm_deger,
                        'otm_kdv_dahil': otm_kdv_dahil,
                        'liste_turu': 'ek2a'
                    })
                    
            except Exception as e:
                print(f"Satır işleme hatası: {str(e)}")
                continue
                
        if not uzmanlik_dallari:
            raise ValueError("İşlenebilir veri bulunamadı")
            
        return uzmanlik_dallari
        
    except Exception as e:
        print(f"Ek2a işleme hatası: {str(e)}")
        return []

def process_ek2b(df):
    """Ek2b Excel dosyasını işler"""
    try:
        # Başlık satırını bul (ilk 20 satıra bak)
        header_row = None
        for i in range(min(20, len(df))):
            row = df.iloc[i]
            # Debug için satırı yazdır
            print(f"{i}. satır str: {[str(x) for x in row if pd.notna(x)]}")
            
            # Boş olmayan değerleri al
            values = [str(x).strip() for x in row if pd.notna(x)]
            # Debug için values'u yazdır
            print(f"{i}. satır values: {values}")
            
            # "İŞLEM KODU" ve "İŞLEM PUANI" kelimeleri geçen satırı bul
            if any("İŞLEM KODU" in str(x).upper() for x in values) and any("İŞLEM PUANI" in str(x).upper() for x in values):
                header_row = i
                break
        
        if header_row is None:
            raise ValueError("Başlık satırı bulunamadı")
        
        # Kolonları belirle
        df.columns = df.iloc[header_row]
        df = df.iloc[header_row + 1:]
        
        # Kolon isimlerini bul
        kod_kolonu = None
        puan_kolonu = None
        
        for col in df.columns:
            if not isinstance(col, str):
                continue
                
            col_upper = str(col).upper().strip()
            if "İŞLEM KODU" in col_upper:
                kod_kolonu = col
            elif "İŞLEM PUANI" in col_upper:
                puan_kolonu = col
        
        if kod_kolonu is None or puan_kolonu is None:
            raise ValueError(f"Gerekli kolonlar bulunamadı. Bulunan kolonlar: {list(df.columns)}")
        
        # Veriyi işle
        islemler = []
        for _, row in df.iterrows():
            try:
                kod = str(row[kod_kolonu]).strip() if pd.notna(row[kod_kolonu]) else ''
                puan = row[puan_kolonu] if pd.notna(row[puan_kolonu]) else None
                
                # Boş satırları ve geçersiz kodları atla
                if not kod or not isinstance(kod, str):
                    continue
                
                # Kod temizleme
                # 1. Başta harf varsa koru
                # 2. Sayısal kısmı bul
                # 3. Sonraki karakterleri at
                kod_temiz = ''
                sayisal_kisim_basladi = False
                
                for c in kod:
                    if c.isdigit():
                        sayisal_kisim_basladi = True
                        kod_temiz += c
                    elif not sayisal_kisim_basladi and c.isalpha():
                        # Sayısal kısım başlamadan önceki harfleri koru
                        kod_temiz += c
                    elif sayisal_kisim_basladi:
                        # Sayısal kısımdan sonraki karakterleri yoksay
                        break
                
                if not kod_temiz or not any(c.isdigit() for c in kod_temiz):
                    continue
                
                # Sayısal değer kontrolü
                if isinstance(puan, str):
                    puan = puan.replace(',', '.').strip()
                    if not puan.replace('.', '').isdigit():
                        continue
                    puan = float(puan)
                elif not isinstance(puan, (int, float)) or pd.isna(puan):
                    continue
                
                # Fiyatları hesapla
                kdv_haric = round(float(puan) * 0.593, 2)
                kdv_dahil = round(kdv_haric * 1.1, 4)
                
                islemler.append({
                    'islem_kodu': kod_temiz,
                    'kdv_haric_fiyat': kdv_haric,
                    'kdv_dahil_fiyat': kdv_dahil,
                    'liste_turu': 'ek2b'
                })
            except (ValueError, TypeError) as e:
                print(f"Satır işleme hatası: {str(e)}")
                continue
                
        if not islemler:
            raise ValueError("İşlenebilir veri bulunamadı")
                
        return islemler
    except Exception as e:
        print(f"Ek2b işleme hatası: {str(e)}")
        return []

def process_ek2c(df):
    """Ek2c Excel dosyasını işler"""
    try:
        # Başlık satırını bul (ilk 20 satıra bak)
        header_row = None
        for i in range(min(20, len(df))):
            row = df.iloc[i]
            # Debug için satırı yazdır
            print(f"{i}. satır str: {[str(x) for x in row if pd.notna(x)]}")
            
            # Boş olmayan değerleri al
            values = [str(x).strip() for x in row if pd.notna(x)]
            # Debug için values'u yazdır
            print(f"{i}. satır values: {values}")
            
            # "İŞLEM KODU" ve "İŞLEM PUANI" kelimeleri geçen satırı bul
            if any("İŞLEM KODU" in str(x).upper() for x in values) and any("İŞLEM PUANI" in str(x).upper() for x in values):
                header_row = i
                break
        
        if header_row is None:
            raise ValueError("Başlık satırı bulunamadı")
        
        # Kolonları belirle
        df.columns = df.iloc[header_row]
        df = df.iloc[header_row + 1:]
        
        # Kolon isimlerini bul
        kod_kolonu = None
        puan_kolonu = None
        
        for col in df.columns:
            if not isinstance(col, str):
                continue
                
            col_upper = str(col).upper().strip()
            if "İŞLEM KODU" in col_upper:
                kod_kolonu = col
            elif "İŞLEM PUANI" in col_upper:
                puan_kolonu = col
        
        if kod_kolonu is None or puan_kolonu is None:
            raise ValueError(f"Gerekli kolonlar bulunamadı. Bulunan kolonlar: {list(df.columns)}")
        
        # Veriyi işle
        islemler = []
        for _, row in df.iterrows():
            try:
                kod = str(row[kod_kolonu]).strip() if pd.notna(row[kod_kolonu]) else ''
                puan = row[puan_kolonu] if pd.notna(row[puan_kolonu]) else None
                
                # Boş satırları ve geçersiz kodları atla
                if not kod or not isinstance(kod, str):
                    continue
                
                # Kod temizleme
                # 1. Başta harf varsa koru
                # 2. Sayısal kısmı bul
                # 3. Sonraki karakterleri at
                kod_temiz = ''
                sayisal_kisim_basladi = False
                
                for c in kod:
                    if c.isdigit():
                        sayisal_kisim_basladi = True
                        kod_temiz += c
                    elif not sayisal_kisim_basladi and c.isalpha():
                        # Sayısal kısım başlamadan önceki harfleri koru
                        kod_temiz += c
                    elif sayisal_kisim_basladi:
                        # Sayısal kısımdan sonraki karakterleri yoksay
                        break
                
                if not kod_temiz or not any(c.isdigit() for c in kod_temiz):
                    continue
                
                # Sayısal değer kontrolü
                if isinstance(puan, str):
                    puan = puan.replace(',', '.').strip()
                    if not puan.replace('.', '').isdigit():
                        continue
                    puan = float(puan)
                elif not isinstance(puan, (int, float)) or pd.isna(puan):
                    continue
                
                # Fiyatları hesapla
                kdv_haric = round(float(puan) * 0.593, 2)
                kdv_dahil = round(kdv_haric * 1.1, 4)
                
                islemler.append({
                    'islem_kodu': kod_temiz,
                    'kdv_haric_fiyat': kdv_haric,
                    'kdv_dahil_fiyat': kdv_dahil,
                    'liste_turu': 'ek2c'
                })
            except (ValueError, TypeError) as e:
                print(f"Satır işleme hatası: {str(e)}")
                continue
                
        if not islemler:
            raise ValueError("İşlenebilir veri bulunamadı")
                
        return islemler
    except Exception as e:
        print(f"Ek2c işleme hatası: {str(e)}")
        return []

# Routes
@app.route('/admin', methods=['GET', 'POST'])
def admin():
    form = AdminForm()
    if form.validate_on_submit():
        if form.clear_data.data:  # Temizle butonuna basıldıysa
            try:
                # Boş JSON yapısını oluştur
                empty_data = {
                    'ek2a': {},
                    'ek2b': {},
                    'ek2c': {}
                }
                # JSON dosyasını temizle
                save_to_json(empty_data)
                flash('Tüm veriler başarıyla temizlendi.', 'success')
            except Exception as e:
                flash(f'Veriler temizlenirken bir hata oluştu: {str(e)}', 'error')
            return redirect(url_for('admin'))
            
        if form.sut_file.data:
            # Dosya seçildi mi kontrol et
            if not form.sut_file.data.filename:
                flash('Lütfen bir dosya seçin.', 'error')
                return redirect(url_for('admin'))
            
            # Dosya adını kontrol et
            filename = form.sut_file.data.filename
            if not (filename.startswith('EK-2A') or filename.startswith('EK-2B') or filename.startswith('EK-2C')):
                flash('Dosya adı EK-2A, EK-2B veya EK-2C ile başlamalıdır.', 'error')
                return redirect(url_for('admin'))
            
            try:
                # Geçici dosyayı kaydet
                temp_path = os.path.join(app.config['UPLOAD_FOLDER'], secure_filename(filename))
                form.sut_file.data.save(temp_path)
                
                # Excel dosyasını oku
                df = pd.read_excel(temp_path)
                
                # Debug için yazdır
                print("\nKolon isimleri:", df.columns.tolist())
                print("\n3. satır (index=2):", df.iloc[2].tolist())
                print("\n3. satır str:", [str(x) for x in df.iloc[2].tolist()])
                
                # Dosya türüne göre işle
                if filename.startswith('EK-2A'):
                    data = process_ek2a(df)
                elif filename.startswith('EK-2B'):
                    data = process_ek2b(df)
                else:  # EK-2C
                    data = process_ek2c(df)
                
                if data:
                    # JSON dosyasını oku
                    json_data = read_and_process_sut_files()
                    if json_data is None:
                        json_data = {'ek2a': {}, 'ek2b': {}, 'ek2c': {}}
                    
                    # Güncel tarihi al
                    current_date = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                    
                    # Dosya türüne göre güncelle
                    if filename.startswith('EK-2A'):
                        json_data['ek2a'][current_date] = data
                    elif filename.startswith('EK-2B'):
                        json_data['ek2b'][current_date] = data
                    else:  # EK-2C
                        json_data['ek2c'][current_date] = data
                    
                    # JSON'a kaydet
                    save_to_json(json_data)
                    flash('Veriler başarıyla güncellendi.', 'success')
                else:
                    flash('Veri işlenirken bir hata oluştu.', 'error')
                
                # Geçici dosyayı sil
                if os.path.exists(temp_path):
                    os.remove(temp_path)
                
            except Exception as e:
                flash(f'Hata: {str(e)}', 'error')
                # Hata durumunda da geçici dosyayı silmeye çalış
                if os.path.exists(temp_path):
                    os.remove(temp_path)
            
            return redirect(url_for('admin'))
    
    return render_template('admin.html', form=form)

@app.route('/', methods=['GET', 'POST'])
def index():
    form = UpdateForm()
    
    if form.validate_on_submit():
        try:
            # Form verilerini al
            excel_file = form.excel_file.data
            hospital_type = form.hospital_type.data
            code_column = form.code_column.data
            description_column = form.description_column.data
            price_column = form.price_column.data
            price_type = form.price_type.data
            
            # Dosya erişilebilirliğini kontrol et
            if not check_file_access(excel_file):
                flash('Excel dosyası açık! Lütfen dosyayı kapatıp tekrar deneyin.', 'error')
                return redirect(url_for('index'))
            
            # Excel dosyasını oku ve güncelle
            result = update_excel_from_json(
                excel_file,
                hospital_type,
                code_column,
                description_column,
                price_column,
                price_type
            )
            
            if result is None or result[0] is None:
                flash('Güncelleme sırasında bir hata oluştu!', 'error')
                return redirect(url_for('index'))
                
            excel_buffer, updated_rows, not_found_rows = result
            
            # Response'u hazırla
            response = send_file(
                excel_buffer,
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                as_attachment=True,
                download_name=os.path.splitext(excel_file.filename)[0] + '_guncel.xlsx'
            )
            
            # Access-Control-Expose-Headers ekle
            response.headers['Access-Control-Expose-Headers'] = 'Content-Disposition, X-Updated-Rows, X-Total-Rows, X-Not-Found-Rows'
            
            # İstatistik bilgilerini header'a ekle
            response.headers['X-Updated-Rows'] = str(updated_rows)
            response.headers['X-Total-Rows'] = str(updated_rows + not_found_rows)
            response.headers['X-Not-Found-Rows'] = str(not_found_rows)
            
            return response
            
        except Exception as e:
            print(f"Hata oluştu: {str(e)}")
            import traceback
            traceback.print_exc()
            flash('Bir hata oluştu!', 'error')
            return redirect(url_for('index'))
            
    return render_template('index.html', form=form)

if __name__ == '__main__':
    app.run(debug=True)
